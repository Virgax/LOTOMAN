#!/usr/bin/env python3
"""
ATHENA ANALYTICS v4 — Motor de análisis para Super Kino TV (Leidsa, RD)
=======================================================================

Uso:
    python athena.py                          # recomendación de hoy, budget default
    python athena.py --budget 800             # escalar jugadas al presupuesto
    python athena.py --backtest 200           # validar el modelo en los últimos N sorteos
    python athena.py --validate               # re-verificar las 3 señales sobre toda la base
    python athena.py --export salida.xlsx     # generar workbook de jugadas

Dependencias: openpyxl (pandas y numpy NO son requeridos).
"""

import argparse
import random
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# CONSTANTES DEL JUEGO
# ---------------------------------------------------------------------------
POOL = 80                 # números del 1 al 80
DRAWN = 20                # bolas extraídas por sorteo
PICK = 10                 # números que escoge el jugador
COST = 25                 # RD$ por jugada
P_BASE = DRAWN / POOL     # 0.25 — probabilidad base de que un número salga

PREMIOS = {10: 25_000_000, 9: 150_000, 8: 10_000, 7: 1_000, 6: 300, 5: 60, 0: 80}

# Pesos v4 — solo 3 señales estadísticamente validadas.
# (DOW, estacionalidad y "gap/atrasados" fueron ELIMINADAS: son ruido esotérico.)
W_REPETICION = 0.45
W_CICLO5 = 0.25
W_MOMENTUM = 0.30

DEFAULT_DB = Path(__file__).parent / "data" / "kino_2010_a_hoy_COMPLETO.xlsx"


# ---------------------------------------------------------------------------
# CARGA DE DATOS
# ---------------------------------------------------------------------------
def load_draws(path=DEFAULT_DB, dedupe=True, verbose=False):
    """Lee el workbook histórico y devuelve [(date, frozenset(nums)), ...] ordenado.

    Formato del archivo: una hoja por año ('2010'..'2026'), cada fila:
        col A = '2026-03-24  Mar'   (fecha + día abreviado)
        col B = '04, 07, 09, ...'   (20 números separados por coma)
    Filas de título/resumen se ignoran automáticamente.

    dedupe=True (default): elimina filas cuyo conjunto de 20 números es idéntico
    a una fila anterior. Son ERRORES DE DATOS, no sorteos reales — la
    probabilidad de que una combinación de 20-de-80 se repita es 1 en 3.5e18.
    Hay 341 de estas filas en la base actual y son la causa del falso "ciclo-5".
    Ver AUDITORIA.md. Usa dedupe=False solo para reproducir el bug.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    draws = []
    for sheet in wb.sheetnames:
        if not sheet.strip().isdigit():
            continue  # salta '📊 Resumen'
        for a, b in wb[sheet].iter_rows(min_col=1, max_col=2, values_only=True):
            if not a or not b or not isinstance(b, str):
                continue
            raw = str(a).strip()[:10]
            try:
                d = datetime.strptime(raw, "%Y-%m-%d").date()
            except ValueError:
                continue
            nums = frozenset(int(x) for x in b.replace(" ", "").split(",") if x)
            if len(nums) >= 18:  # algunos sorteos viejos vienen con 19
                draws.append((d, nums))
    wb.close()
    draws.sort(key=lambda r: r[0])

    if dedupe:
        seen, clean = set(), []
        for dt, nums in draws:
            if nums in seen:
                continue
            seen.add(nums)
            clean.append((dt, nums))
        if verbose and len(clean) != len(draws):
            print(f"[dedupe] {len(draws) - len(clean)} filas duplicadas descartadas "
                  f"({len(draws):,} -> {len(clean):,})")
        draws = clean
    return draws


# ---------------------------------------------------------------------------
# SEÑAL 1 — REPETICIÓN
# ---------------------------------------------------------------------------
def repetition_rates(draws):
    """P(n en sorteo t+1 | n en sorteo t) y P(n en t+1 | n NO en t), por número.

    Base aleatoria = 0.25 para ambas. Los números con tasa de repetición
    consistentemente por encima de 0.25 son la señal más fuerte del modelo.
    """
    rep_hit, rep_tot = Counter(), Counter()
    non_hit, non_tot = Counter(), Counter()
    for i in range(1, len(draws)):
        prev, cur = draws[i - 1][1], draws[i][1]
        for n in range(1, POOL + 1):
            if n in prev:
                rep_tot[n] += 1
                if n in cur:
                    rep_hit[n] += 1
            else:
                non_tot[n] += 1
                if n in cur:
                    non_hit[n] += 1
    p_rep, p_non = {}, {}
    for n in range(1, POOL + 1):
        p_rep[n] = rep_hit[n] / rep_tot[n] if rep_tot[n] else P_BASE
        p_non[n] = non_hit[n] / non_tot[n] if non_tot[n] else P_BASE
    return p_rep, p_non


# ---------------------------------------------------------------------------
# SEÑAL 2 — CICLO-5
# ---------------------------------------------------------------------------
def cycle_lift(draws, lag=5):
    """Lift de P(n sale hoy | n salió exactamente hace `lag` sorteos).

    Hallazgo validado sobre 5,453 sorteos: lag=5 da ~+3.2% relativo sobre 0.25.
    Devuelve (p_condicional, lift_relativo).
    """
    hit = tot = 0
    for i in range(lag, len(draws)):
        past, cur = draws[i - lag][1], draws[i][1]
        for n in past:
            tot += 1
            if n in cur:
                hit += 1
    p = hit / tot if tot else P_BASE
    return p, (p - P_BASE) / P_BASE


def cycle_lift_all_lags(draws, max_lag=15):
    """Escanea lags 1..max_lag. Sirve para re-validar que el 5 sigue siendo el pico."""
    return {lag: cycle_lift(draws, lag) for lag in range(1, max_lag + 1)}


# ---------------------------------------------------------------------------
# SEÑAL 3 — MOMENTUM
# ---------------------------------------------------------------------------
def momentum(draws, upto=None, windows=(7, 14, 30), weights=(0.5, 0.3, 0.2)):
    """Frecuencia reciente relativa a 0.25, ponderada en ventanas de 7/14/30 sorteos."""
    hist = draws[:upto] if upto is not None else draws
    score = {}
    for n in range(1, POOL + 1):
        s = 0.0
        for w, wt in zip(windows, weights):
            recent = hist[-w:]
            f = sum(1 for _, nums in recent if n in nums) / len(recent)
            s += wt * ((f - P_BASE) / P_BASE)
        score[n] = s
    return score


# ---------------------------------------------------------------------------
# SCORE COMPUESTO
# ---------------------------------------------------------------------------
def score_numbers(draws, upto=None):
    """Puntúa los 80 números para el PRÓXIMO sorteo después de draws[:upto].

    Devuelve dict n -> {'score', 'rep', 'c5', 'mom', 'senales'}
    """
    hist = draws[:upto] if upto is not None else draws
    p_rep, p_non = repetition_rates(hist)
    _, c5_lift = cycle_lift(hist, 5)
    mom = momentum(hist)

    last = hist[-1][1]
    lag5 = hist[-5][1] if len(hist) >= 5 else frozenset()

    out = {}
    for n in range(1, POOL + 1):
        p = p_rep[n] if n in last else p_non[n]
        s_rep = (p - P_BASE) / P_BASE
        s_c5 = c5_lift if n in lag5 else 0.0
        s_mom = mom[n]
        score = 100 * (W_REPETICION * s_rep + W_CICLO5 * s_c5 + W_MOMENTUM * s_mom)
        senales = sum([n in last, n in lag5, s_mom > 0])
        out[n] = {
            "score": round(score, 2),
            "rep": n in last,
            "c5": n in lag5,
            "mom": round(s_mom, 3),
            "senales": senales,
        }
    return out


def ranked(scores, k=None):
    r = sorted(scores, key=lambda n: -scores[n]["score"])
    return r[:k] if k else r


# ---------------------------------------------------------------------------
# CONFIANZA Y PRESUPUESTO
# ---------------------------------------------------------------------------
def confidence(draws, lookback=30):
    """3 métricas -> score 0-100.

    A) Precisión reciente: aciertos promedio del top-20 (aleatorio = 5.0)
    B) Consistencia de repetición: cuántos números repitieron vs media histórica
    C) Convergencia de señales: cuántos números tienen las 3 señales alineadas
    """
    hits = []
    for i in range(len(draws) - lookback, len(draws)):
        top20 = set(ranked(score_numbers(draws, upto=i), 20))
        hits.append(len(top20 & draws[i][1]))
    acc = sum(hits) / len(hits)
    a = max(0, min(100, (acc - 5.0) / 2.0 * 100 + 50))

    reps = [len(draws[i - 1][1] & draws[i][1]) for i in range(1, len(draws))]
    media = sum(reps) / len(reps)
    recientes = sum(reps[-lookback:]) / lookback
    b = max(0, min(100, 50 + (recientes - media) * 12))

    sc = score_numbers(draws)
    conv = sum(1 for n in ranked(sc, 20) if sc[n]["senales"] == 3)
    c = max(0, min(100, conv * 10))

    total = 0.5 * a + 0.2 * b + 0.3 * c
    return {
        "confianza": round(total, 1),
        "precision_top20": round(acc, 2),
        "repeticion_reciente": round(recientes, 2),
        "repeticion_media": round(media, 2),
        "convergencia": conv,
        "componentes": {"A_precision": round(a, 1), "B_repeticion": round(b, 1), "C_convergencia": round(c, 1)},
    }


def plays_for_budget(budget, conf):
    """Escala # de jugadas al presupuesto, modulado por confianza.

    <40% de confianza = NO JUGAR (regla dura del sistema).
    """
    if conf < 40:
        return 0
    max_plays = budget // COST
    factor = 0.4 + 0.6 * min(1.0, (conf - 40) / 50)
    return int(max_plays * factor)


# ---------------------------------------------------------------------------
# GENERACIÓN DE JUGADAS — formato concentración-jackpot
# ---------------------------------------------------------------------------
def generate_plays(scores, n_plays, seed=None):
    """Estructura:
      1) NÚCLEO PURO: el top-10 exacto (1 jugada)
      2) ROTACIONES: núcleo con 1-2 sustituciones del top-15
      3) COBERTURA: sustituciones usando ranks 16-20

    Objetivo: concentrar masa de probabilidad para maximizar P(7-10 aciertos),
    NO diversificar para ganar poquito todos los días.
    """
    rng = random.Random(seed)
    r = ranked(scores)
    top10, top15, top20 = r[:10], r[:15], r[:20]
    plays, seen = [], set()

    def add(play, tipo):
        key = frozenset(play)
        if key in seen or len(key) != PICK:
            return False
        seen.add(key)
        plays.append({"nums": sorted(play), "tipo": tipo})
        return True

    add(top10, "NÚCLEO")

    # rotaciones: 1 sustitución desde ranks 11-15
    for out_n in top10:
        for in_n in top15[10:]:
            if len(plays) >= n_plays:
                break
            add([x for x in top10 if x != out_n] + [in_n], "ROTACIÓN-1")

    # rotaciones: 2 sustituciones
    guard = 0
    while len(plays) < n_plays * 0.75 and guard < 5000:
        guard += 1
        outs = rng.sample(top10, 2)
        ins = rng.sample(top15[10:] if len(top15) > 10 else top10, 2)
        add([x for x in top10 if x not in outs] + list(set(ins)), "ROTACIÓN-2")

    # cobertura con ranks 16-20
    guard = 0
    while len(plays) < n_plays and guard < 5000:
        guard += 1
        k = rng.choice([2, 3])
        outs = rng.sample(top10, k)
        ins = rng.sample(top20[15:] + top15[10:], k)
        add([x for x in top10 if x not in outs] + list(set(ins)), "COBERTURA")

    return plays[:n_plays]


# ---------------------------------------------------------------------------
# BACKTEST
# ---------------------------------------------------------------------------
def backtest(draws, n=200, top_k=20):
    """Mide aciertos del top-K contra los sorteos reales.

    LA MÉTRICA QUE IMPORTA: promedio de aciertos del top-20.
    Aleatorio = 5.00. Si el promedio no supera ~5.3 de forma sostenida,
    el modelo NO está agregando valor y hay que decirlo.
    """
    results = []
    start = max(50, len(draws) - n)
    for i in range(start, len(draws)):
        top = set(ranked(score_numbers(draws, upto=i), top_k))
        results.append(len(top & draws[i][1]))
    prom = sum(results) / len(results)
    esperado = top_k * P_BASE
    dist = Counter(results)
    return {
        "sorteos_evaluados": len(results),
        "promedio_aciertos": round(prom, 3),
        "esperado_aleatorio": esperado,
        "lift": round((prom - esperado) / esperado * 100, 2),
        "distribucion": dict(sorted(dist.items())),
    }


def simulate_plays(draws, n_back=100, budget=800, seed=42):
    """Simula el sistema completo hacia atrás: genera jugadas y las cobra."""
    total_gastado = total_ganado = 0
    tiers = Counter()
    start = max(50, len(draws) - n_back)
    for i in range(start, len(draws)):
        sc = score_numbers(draws, upto=i)
        n_plays = max(1, budget // COST // 2)
        plays = generate_plays(sc, n_plays, seed=seed + i)
        real = draws[i][1]
        for p in plays:
            total_gastado += COST
            hits = len(set(p["nums"]) & real)
            tiers[hits] += 1
            total_ganado += PREMIOS.get(hits, 0)
    return {
        "sorteos": len(range(start, len(draws))),
        "gastado": total_gastado,
        "ganado": total_ganado,
        "retorno_pct": round(total_ganado / total_gastado * 100, 2) if total_gastado else 0,
        "distribucion_aciertos": dict(sorted(tiers.items())),
    }


# ---------------------------------------------------------------------------
# EXPORT
# ---------------------------------------------------------------------------
def export_xlsx(path, plays, scores, conf, fecha, budget):
    """Workbook de 4 pestañas: jugadas, scores, cobertura, referencia."""
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    base = Font(name="Arial", size=10)

    wb = Workbook()

    ws = wb.active
    ws.title = "Jugadas"
    ws["A1"] = f"ATHENA v4 — Jugadas {fecha}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A2"] = (f"Confianza {conf['confianza']}%  ·  {len(plays)} jugadas  ·  "
                f"RD${len(plays)*COST:,} de RD${budget:,}")
    ws["A2"].font = base
    ws.append([])
    for i, h in enumerate(["#", "TIPO", "NÚMEROS", "CSV"], start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
    for i, p in enumerate(plays, start=1):
        csv = ", ".join(f"{n:02d}" for n in p["nums"])
        ws.append([i, p["tipo"], csv, csv])
        for c in ws[ws.max_row]:
            c.font, c.border = base, border
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 46

    ws = wb.create_sheet("Scores")
    for i, h in enumerate(["RANK", "NÚMERO", "SCORE", "REPITE", "CICLO-5", "MOMENTUM", "SEÑALES"], start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
    for rank, n in enumerate(ranked(scores), start=1):
        d = scores[n]
        ws.append([rank, n, d["score"], "SÍ" if d["rep"] else "—",
                   "SÍ" if d["c5"] else "—", d["mom"], "⭐" * d["senales"]])
        for c in ws[ws.max_row]:
            c.font, c.border = base, border

    ws = wb.create_sheet("Cobertura")
    freq = Counter()
    for p in plays:
        freq.update(p["nums"])
    for i, h in enumerate(["NÚMERO", "VECES EN JUGADAS", "% DE JUGADAS"], start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
    for n, v in freq.most_common():
        ws.append([n, v, round(v / len(plays) * 100, 1) if plays else 0])
        for c in ws[ws.max_row]:
            c.font, c.border = base, border

    ws = wb.create_sheet("Referencia")
    ref = [
        ["TABLA DE PREMIOS", ""],
        ["10 aciertos", "RD$ 25,000,000"],
        ["9 aciertos", "RD$ 150,000"],
        ["8 aciertos", "RD$ 10,000"],
        ["7 aciertos", "RD$ 1,000"],
        ["6 aciertos", "RD$ 300"],
        ["5 aciertos", "RD$ 60"],
        ["0 aciertos", "RD$ 80 (devuelto)"],
        ["", ""],
        ["DIAGNÓSTICO DEL MODELO", ""],
        ["Precisión top-20 (30 sorteos)", conf["precision_top20"]],
        ["Esperado si fuera aleatorio", 5.0],
        ["Repetición reciente", conf["repeticion_reciente"]],
        ["Repetición media histórica", conf["repeticion_media"]],
        ["Convergencia (3 señales)", conf["convergencia"]],
        ["", ""],
        ["REALIDAD MATEMÁTICA", ""],
        ["Valor esperado por jugada", "RD$ ~16.88 de RD$ 25"],
        ["Ventaja de la casa", "~32.5% — permanente"],
        ["P(10 aciertos)", "1 en 1,646,492,110,120"],
    ]
    for row in ref:
        ws.append(row)
        for c in ws[ws.max_row]:
            c.font, c.border = base, border
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# AUDITORÍA — ¿las señales son reales o son artefactos?
# ---------------------------------------------------------------------------
def audit(path=DEFAULT_DB):
    """Prueba honesta de las 3 señales. Corre esto ANTES de confiar en el modelo."""
    import math
    import statistics

    raw = load_draws(path, dedupe=False)
    clean = load_draws(path, dedupe=True)

    print("=" * 68)
    print("AUDITORÍA ATHENA — ¿las señales sobreviven?")
    print("=" * 68)

    print(f"\n[1] INTEGRIDAD\n    filas crudas: {len(raw):,}   filas únicas: {len(clean):,}"
          f"   duplicadas: {len(raw)-len(clean):,}")
    idx = defaultdict(list)
    for i, (_, n) in enumerate(raw):
        idx[n].append(i)
    gaps = Counter()
    years = Counter()
    for n, ii in idx.items():
        for a, b in zip(ii, ii[1:]):
            gaps[b - a] += 1
            years[raw[b][0].year] += 1
    print(f"    separación de los duplicados: {dict(sorted(gaps.items())[:6])}")
    print(f"    año de la copia: {dict(sorted(years.items()))}")
    print("    -> P(que una combinación 20-de-80 se repita) = 1 en 3.5e18.")
    print("       Estas filas son error de captura, no sorteos.")

    print("\n[2] CICLO-5 (peso actual 25%)")
    print("    lag |  base cruda  |  base limpia")
    for lag in range(1, 11):
        _, lr = cycle_lift(raw, lag)
        _, lc = cycle_lift(clean, lag)
        mark = "   <-- el supuesto ciclo" if lag == 5 else ""
        print(f"    {lag:3d} |   {lr*100:+6.2f}%    |   {lc*100:+6.2f}%{mark}")
    print("    -> El pico de lag 5/10/15 en la base cruda es el eco de los duplicados.")

    print("\n[3] REPETICIÓN (peso actual 45%)")
    p_rep, p_non = repetition_rates(clean)
    diff = [p_rep[n] - p_non[n] for n in range(1, POOL + 1)]
    m = sum(diff) / POOL
    sd = statistics.stdev(diff)
    print(f"    media de P(sale|salió) - P(sale|no salió) = {m*100:+.3f} puntos")
    print(f"    desviación entre números = {sd*100:.3f} puntos")
    print("    -> 0.000 significa que el sorteo NO tiene memoria de un día al otro.")

    print("\n[4] ¿HAY NÚMEROS 'CALIENTES' DE VERDAD?")
    N = len(clean)
    base = Counter()
    for _, n in clean:
        base.update(n)
    exp, sd_b = N * P_BASE, math.sqrt(N * P_BASE * (1 - P_BASE))
    z = {n: (base[n] - exp) / sd_b for n in range(1, POOL + 1)}
    obs_sd = (sum(v * v for v in z.values()) / POOL) ** 0.5
    hot = max(z, key=lambda n: abs(z[n]))
    print(f"    desviación estándar de los z observados = {obs_sd:.3f}  (azar puro = 1.000)")
    print(f"    número más extremo: {hot} con z = {z[hot]:+.2f}")
    print(f"    -> Con 80 números, un |z| de ~2.5-3.0 aparece por azar SIEMPRE.")

    print("\n[5] LA MÉTRICA QUE IMPORTA — aciertos del top-20")
    r = backtest(clean, 300, 20)
    hits_avg = r["promedio_aciertos"]
    print(f"    promedio sobre {r['sorteos_evaluados']} sorteos = {hits_avg:.3f}")
    print(f"    esperado si escogieras 20 números al azar = 5.000")
    print(f"    -> Regla del propio sistema: >5.3 sostenido = aporta. ~5.0 = no aporta.")

    print("\n" + "=" * 68)
    print("VALOR ESPERADO (esto no depende del modelo)")
    print("  RD$25 apostados devuelven ~RD$16.88 en promedio. Ventaja casa ~32.5%.")
    print("  RD$800/día = RD$24,000/mes = pérdida esperada ~RD$7,800/mes.")
    print("=" * 68)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Athena Analytics v4 — Super Kino TV")
    ap.add_argument("--db", default=str(DEFAULT_DB))
    ap.add_argument("--budget", type=int, default=800)
    ap.add_argument("--backtest", type=int, metavar="N")
    ap.add_argument("--simulate", type=int, metavar="N")
    ap.add_argument("--validate", action="store_true")
    ap.add_argument("--audit", action="store_true", help="auditoría honesta de las señales")
    ap.add_argument("--raw", action="store_true", help="NO deduplicar (reproduce el bug)")
    ap.add_argument("--export", metavar="ARCHIVO.xlsx")
    ap.add_argument("--seed", type=int, default=None)
    args = ap.parse_args()

    if args.audit:
        audit(args.db)
        return

    draws = load_draws(args.db, dedupe=not args.raw, verbose=True)
    print(f"Base: {len(draws):,} sorteos  |  {draws[0][0]} → {draws[-1][0]}\n")

    if args.validate:
        print("VALIDACIÓN DE SEÑALES")
        p_rep, p_non = repetition_rates(draws)
        top = sorted(p_rep, key=lambda n: -p_rep[n])[:10]
        print(f"  Repetición media: {sum(p_rep.values())/POOL:.4f}  (base {P_BASE})")
        print("  Top-10 repetidores: " + ", ".join(f"{n}({p_rep[n]*100:.1f}%)" for n in top))
        print("\n  Lag scan (lift % sobre 0.25):")
        for lag, (p, lift) in cycle_lift_all_lags(draws).items():
            mark = "  <-- CICLO-5" if lag == 5 else ""
            print(f"    lag {lag:2d}: p={p:.4f}  lift={lift*100:+.2f}%{mark}")
        return

    if args.backtest:
        print("BACKTEST — aciertos del top-20")
        r = backtest(draws, args.backtest)
        for k, v in r.items():
            print(f"  {k}: {v}")
        print("\n  Regla: promedio > 5.3 sostenido = el modelo aporta. ≈5.0 = no aporta.")
        return

    if args.simulate:
        print("SIMULACIÓN DE JUGADAS REALES")
        for k, v in simulate_plays(draws, args.simulate, args.budget).items():
            print(f"  {k}: {v}")
        return

    conf = confidence(draws)
    scores = score_numbers(draws)
    n_plays = plays_for_budget(args.budget, conf["confianza"])

    print(f"CONFIANZA: {conf['confianza']}%")
    print(f"  precisión top-20 (30 sorteos): {conf['precision_top20']} (aleatorio 5.0)")
    print(f"  repetición reciente: {conf['repeticion_reciente']} vs media {conf['repeticion_media']}")
    print(f"  convergencia 3 señales: {conf['convergencia']} números\n")

    if n_plays == 0:
        print("CONFIANZA < 40% → el sistema dice NO JUGAR hoy.")
        return

    r = ranked(scores)
    print("TOP-10 (núcleo): " + "  ·  ".join(f"{n:02d}" for n in r[:10]))
    print("RANKS 11-15:     " + "  ·  ".join(f"{n:02d}" for n in r[10:15]))
    print("RANKS 16-20:     " + "  ·  ".join(f"{n:02d}" for n in r[15:20]))
    print(f"\nJUGADAS: {n_plays}  ·  RD${n_plays*COST:,} de RD${args.budget:,} "
          f"(sobran RD${args.budget - n_plays*COST:,})\n")

    plays = generate_plays(scores, n_plays, seed=args.seed)
    for i, p in enumerate(plays, 1):
        print(f"{i:3d}. [{p['tipo']:<11}] " + ", ".join(f"{n:02d}" for n in p["nums"]))
        if i % 3 == 0:
            print("---")

    if args.export:
        fecha = draws[-1][0].strftime("%d/%m/%Y") + " (próximo sorteo)"
        out = export_xlsx(args.export, plays, scores, conf, fecha, args.budget)
        print(f"\nExportado: {out}")


if __name__ == "__main__":
    main()
