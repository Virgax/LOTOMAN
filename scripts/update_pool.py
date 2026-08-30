#!/usr/bin/env python3
"""
Actualizador de la base histórica de Loto Pool (Leidsa, RD).

Uso:
    python scripts/update_pool.py                       # desde el último hasta hoy
    python scripts/update_pool.py --desde 2026-01-01 --hasta 2026-08-26
    python scripts/update_pool.py --dry-run             # ver sin escribir

Fuente: resuloto.com
    https://www.resuloto.com/do/leid/loto-pool-amp.php?fecha=YYYY-MM-DD

EL JUEGO (verificado, no supuesto)
  * 5 números del 00 al 31.
  * Sorteo DIARIO. Esto se comprobó barriendo 9 días corridos (2026-08-17 a
    08-25, domingo incluido): los nueve traen resultado. conectate.com se
    contradice a sí mismo — su índice /leidsa/ lo lista como diario y su
    página de Loto Pool dice "miércoles y sábados". El dato manda: es diario.

DIFERENCIA CRÍTICA CON KINO — no copiar la regla anti-duplicados
  update_db.py rechaza cualquier sorteo cuya combinación ya exista, porque en
  Kino repetir 20-de-80 tiene probabilidad 1 en 3.5e18: si pasa, es error de
  captura. En Loto Pool NO aplica:

      C(32,5)                                   =   201,376 combinaciones
      repeticiones esperadas en ~5,600 sorteos  =        78

  Setenta y ocho repeticiones legítimas por puro azar. Aplicar aquí la regla
  de Kino borraría 78 sorteos reales. Por eso esta base deduplica por FECHA,
  nunca por números.

EL PARSER va anclado a la estructura, no a posiciones. Los 5 números son los
que preceden al "Anterior" de la navegación:

    Loto Pool Sábado 22 Ago  09 15 22 29 31  ◄ Anterior  Siguiente ►

y la fecha sale del título ("Resultado del Sábado 22 de Agosto de 2026"), no
de la fecha que se pidió: así se detecta si el sitio sirve otro día.
Contar números sueltos en la página NO sirve: ya falló dos veces en este
proyecto y así entraron las 341 filas basura de 2011.
"""

import argparse
import re
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

DB = Path(__file__).parent.parent / "data" / "loto_pool.xlsx"
URL = "https://www.resuloto.com/do/leid/loto-pool-amp.php?fecha={}"
DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; AthenaAnalytics/4.0)"}

POOL_MIN, POOL_MAX, POOL_N = 1, 31, 5   # verificado: el 00 NO sale nunca en 18,600 bolas
MESES = {m: i + 1 for i, m in enumerate(
    "enero febrero marzo abril mayo junio julio agosto "
    "septiembre octubre noviembre diciembre".split())}


def _sin_tildes(s):
    return unicodedata.normalize("NFKD", s.lower()).encode("ascii", "ignore").decode()


def _texto(html):
    h = re.sub(r"<(script|style).*?</\1>", " ", html, flags=re.S | re.I)
    h = re.sub(r"<[^>]+>", " ", h)
    for ent, ch in (("&aacute;", "á"), ("&eacute;", "é"), ("&iacute;", "í"),
                    ("&oacute;", "ó"), ("&uacute;", "ú"), ("&nbsp;", " ")):
        h = h.replace(ent, ch)
    return re.sub(r"\s+", " ", h).strip()


def parse_pool(html):
    """(fecha_del_titulo | None, [5 numeros] | None)."""
    t = _texto(html)
    fecha = None
    m = re.search(r"Resultado del \w+ (\d{1,2}) de ([A-Za-zÁ-ú]+) de (\d{4})", t)
    if m:
        mes = MESES.get(_sin_tildes(m.group(2)))
        if mes:
            try:
                fecha = date(int(m.group(3)), mes, int(m.group(1)))
            except ValueError:
                fecha = None

    n = re.search(r"((?:\b\d{2}\b\s+){%d})(?:&#9668;|◄)?\s*Anterior" % POOL_N, t)
    if not n:
        return fecha, None
    v = [int(x) for x in n.group(1).split()]
    if (len(v) != POOL_N or len(set(v)) != POOL_N
            or not all(POOL_MIN <= x <= POOL_MAX for x in v)):
        return fecha, None
    return fecha, sorted(v)


def leer_base(path=DB):
    """{date: '00, 00, ...'} de lo ya guardado. Vacío si la base no existe."""
    if not Path(path).exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    datos = {}
    for hoja in wb.sheetnames:
        if not hoja.strip().isdigit():
            continue
        for a, b in wb[hoja].iter_rows(min_col=1, max_col=2, values_only=True):
            if not a or not b or not isinstance(b, str):
                continue
            try:
                d = datetime.strptime(str(a).strip()[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
            datos[d] = b.strip()
    wb.close()
    return datos


def scrape(d):
    """Los 5 números de una fecha, o None. Verifica que el sitio sirva ESE día."""
    try:
        r = requests.get(URL.format(d.isoformat()), headers=HEADERS, timeout=25)
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"  ! {d} error de red: {e}", file=sys.stderr)
        return None
    if not r.encoding or r.encoding.lower() in ("iso-8859-1", "latin-1"):
        r.encoding = r.apparent_encoding or "utf-8"

    fecha, nums = parse_pool(r.text)
    if nums is None:
        return None
    if fecha and fecha != d:
        # El sitio devolvió otro día. Guardarlo bajo la fecha pedida sería
        # inventar un sorteo: exactamente el error que ensució la base de Kino.
        print(f"  ! {d} RECHAZADO: la página dice {fecha}", file=sys.stderr)
        return None
    return nums


def escribir(nuevos, path=DB):
    path = Path(path)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
    for d, nums in sorted(nuevos.items()):
        hoja = str(d.year)
        if hoja not in wb.sheetnames:
            ws = wb.create_sheet(hoja)
            ws["A1"] = f"Loto Pool — {hoja}"
            ws["A2"], ws["B2"] = "FECHA", "NÚMEROS GANADORES (5)"
            for c in (ws["A2"], ws["B2"]):
                c.font = Font(name="Arial", bold=True)
        ws = wb[hoja]
        fila = ws.max_row + 1
        ws.cell(row=fila, column=1,
                value=f"{d.isoformat()}  {DIAS[d.weekday()]}").font = Font(name="Arial", size=10)
        ws.cell(row=fila, column=2,
                value=", ".join(f"{n:02d}" for n in nums)).font = Font(name="Arial", size=10)
    # Hojas por año en orden, que si no quedan como salieron.
    for i, nombre in enumerate(sorted(wb.sheetnames)):
        wb.move_sheet(nombre, offset=i - wb.sheetnames.index(nombre))
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=str(DB))
    ap.add_argument("--desde")
    ap.add_argument("--hasta")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--pausa", type=float, default=1.0)
    args = ap.parse_args()

    existente = leer_base(args.db)
    if existente:
        print(f"Base actual: {len(existente):,} sorteos, último = {max(existente)}")
    else:
        print("Base de Loto Pool vacía — se crea desde cero.")

    if args.desde:
        desde = date.fromisoformat(args.desde)
    elif existente:
        desde = max(existente) + timedelta(days=1)
    else:
        desde = date.today() - timedelta(days=30)
    hasta = date.fromisoformat(args.hasta) if args.hasta else date.today()
    if desde > hasta:
        print("Nada que actualizar.")
        return

    nuevos, faltantes = {}, []
    d = desde
    while d <= hasta:
        if d in existente:          # dedupe por FECHA. Por números NO: ver el
            d += timedelta(days=1)  # encabezado, se esperan ~78 repeticiones
            continue                # legítimas en una historia completa.
        nums = scrape(d)
        if nums:
            nuevos[d] = nums
            print(f"  + {d} {DIAS[d.weekday()]}  " + ", ".join(f"{n:02d}" for n in nums))
        else:
            faltantes.append(d)
            print(f"  - {d} sin resultado")
        time.sleep(args.pausa)
        d += timedelta(days=1)

    print(f"\nNuevos: {len(nuevos)}  |  Sin resultado: {len(faltantes)}")
    if faltantes:
        print("  Días sin dato: " + ", ".join(str(x) for x in faltantes[:20]))
        print("  NO los rellenes a mano copiando otro día.")

    # Mismo seguro que update_db.py: varios días pedidos y ninguno con dato es
    # la fuente rota, no una racha de sorteos suspendidos.
    if not nuevos and len(faltantes) > 3:
        print(f"\n! ABORTADO: {len(faltantes)} días pedidos, ninguno trajo resultado.",
              file=sys.stderr)
        print("  Revisa si resuloto.com cambió de formato.", file=sys.stderr)
        sys.exit(2)

    if args.dry_run:
        print("(dry-run: no se escribió nada)")
        return
    if nuevos:
        escribir(nuevos, args.db)
        print(f"Base actualizada: {len(existente) + len(nuevos):,} sorteos")


if __name__ == "__main__":
    main()
