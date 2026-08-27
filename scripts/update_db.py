#!/usr/bin/env python3
"""
Actualizador de la base histórica de Super Kino TV.

Uso:
    python scripts/update_db.py                      # desde el último sorteo hasta hoy
    python scripts/update_db.py --desde 2026-05-25 --hasta 2026-08-26
    python scripts/update_db.py --dry-run            # ver qué traería sin escribir

Fuente: resuloto.com
    https://www.resuloto.com/do/leid/super-kino-tv-amp.php?fecha=YYYY-MM-DD

NOTAS IMPORTANTES (aprendidas a la mala):
  * Leidsa NO sortea Jueves Santo ni Viernes Santo. Un día sin resultado ahí
    es normal, no es un fallo del scraper.
  * NUNCA rellenes un día faltante copiando el resultado del día anterior.
    Eso fue exactamente lo que metió 341 filas basura en la base y creó
    el falso "ciclo-5". Si un día no tiene resultado, se deja fuera.
  * El script rechaza automáticamente cualquier resultado idéntico a otro
    ya presente en la base.

FUENTE ALTERNATIVA (elboletoganador.com):
    api3.bolillerobingoonlinegratis.com/api/sorteos/buscar/historial
    - `fecha` funciona como CURSOR HACIA ATRÁS, no como número de página
    - devuelve 15 resultados por llamada
    - Game IDs: Kino=8, Pool=7, Pega3=23, Quiniela=5
    - requiere inyección de content script en una pestaña abierta de
      elboletoganador.com (las peticiones desde el origen de la extensión
      están bloqueadas por CORS)
    - firma obligatoria: apiFetchViaTab(tabId, gameId, fecha, cb)
      agregar un parámetro `page` ROMPE el callback
"""

import argparse
import re
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

DB = Path(__file__).parent.parent / "data" / "kino_2010_a_hoy_COMPLETO.xlsx"
URL = "https://www.resuloto.com/do/leid/super-kino-tv-amp.php?fecha={}"
DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; AthenaAnalytics/4.0)"}


def leer_base(path=DB):
    """Devuelve {date: 'nn, nn, ...'} de todo lo que ya está guardado."""
    wb = load_workbook(path, read_only=True, data_only=True)
    datos = {}
    for hoja in wb.sheetnames:
        if not hoja.strip().isdigit():
            continue
        for a, b in wb[hoja].iter_rows(min_col=1, max_col=2, values_only=True):
            if not a or not b or not isinstance(b, str):
                continue
            try:
                d = datetime.strptime(str(a).strip()[:10], "%Y-%m-%d").date()
            except ValueError:
                continue
            datos[d] = b.strip()
    wb.close()
    return datos


def scrape(d):
    """Trae los 20 números de una fecha. Devuelve lista de int o None."""
    try:
        r = requests.get(URL.format(d.isoformat()), headers=HEADERS, timeout=20)
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"  ! {d} error de red: {e}", file=sys.stderr)
        return None

    html = re.sub(r"<script.*?</script>", " ", r.text, flags=re.S | re.I)
    texto = re.sub(r"<[^>]+>", " ", html)
    nums = [int(x) for x in re.findall(r"\b(\d{1,2})\b", texto) if 1 <= int(x) <= 80]

    # busca una ventana de 20 números distintos: es el resultado del sorteo
    for i in range(len(nums) - 19):
        v = nums[i:i + 20]
        if len(set(v)) == 20:
            return sorted(set(v))
    return None


def escribir(nuevos, path=DB):
    """Agrega las filas nuevas a la hoja del año correspondiente."""
    wb = load_workbook(path)
    for d, nums in sorted(nuevos.items()):
        hoja = str(d.year)
        if hoja not in wb.sheetnames:
            ws = wb.create_sheet(hoja)
            ws["A1"] = f"Super Kino TV — {hoja}"
            ws["A2"] = "FECHA"
            ws["B2"] = "NÚMEROS GANADORES (20)"
            for c in (ws["A2"], ws["B2"]):
                c.font = Font(name="Arial", bold=True)
        ws = wb[hoja]
        etiqueta = f"{d.isoformat()}  {DIAS[d.weekday()]}"
        cadena = ", ".join(f"{n:02d}" for n in nums)
        fila = ws.max_row + 1
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(name="Arial", size=10)
        ws.cell(row=fila, column=2, value=cadena).font = Font(name="Arial", size=10)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=str(DB))
    ap.add_argument("--desde")
    ap.add_argument("--hasta")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--pausa", type=float, default=1.0, help="segundos entre peticiones")
    args = ap.parse_args()

    existente = leer_base(args.db)
    conjuntos = {v for v in existente.values()}
    ultimo = max(existente)
    print(f"Base actual: {len(existente):,} sorteos, último = {ultimo}")

    desde = date.fromisoformat(args.desde) if args.desde else ultimo + timedelta(days=1)
    hasta = date.fromisoformat(args.hasta) if args.hasta else date.today()
    if desde > hasta:
        print("Nada que actualizar.")
        return

    nuevos, faltantes = {}, []
    d = desde
    while d <= hasta:
        if d in existente:
            d += timedelta(days=1)
            continue
        nums = scrape(d)
        if nums:
            cadena = ", ".join(f"{n:02d}" for n in nums)
            if cadena in conjuntos:
                print(f"  ! {d} RECHAZADO: resultado idéntico a otro ya guardado "
                      f"(esto es lo que ensució la base antes)")
            else:
                nuevos[d] = nums
                conjuntos.add(cadena)
                print(f"  + {d} {DIAS[d.weekday()]}  {cadena}")
        else:
            faltantes.append(d)
            print(f"  - {d} sin resultado (¿Jueves/Viernes Santo? ¿sorteo no realizado?)")
        time.sleep(args.pausa)
        d += timedelta(days=1)

    print(f"\nNuevos: {len(nuevos)}  |  Sin resultado: {len(faltantes)}")
    if faltantes:
        print("  Días sin dato: " + ", ".join(str(x) for x in faltantes))
        print("  NO los rellenes a mano copiando otro día.")
    # Guardia anti-fallo-silencioso. Si se pidieron varios días y NINGUNO trajo
    # resultado, lo más probable es que resuloto.com cambió de formato o está
    # caída — no que Leidsa dejara de sortear una semana entera. Abortamos con
    # error para que la rutina diaria se ponga roja en vez de reportar
    # "sin resultado" en silencio para siempre.
    if not nuevos and len(faltantes) > 3:
        print(f"\n! ABORTADO: {len(faltantes)} días pedidos, ninguno trajo resultado.",
              file=sys.stderr)
        print("  Eso no es un feriado. Revisa si la fuente cambió de formato.",
              file=sys.stderr)
        sys.exit(2)

    if args.dry_run:
        print("(dry-run: no se escribió nada)")
        return
    if nuevos:
        escribir(nuevos, args.db)
        print(f"Base actualizada: {len(existente)+len(nuevos):,} sorteos")


if __name__ == "__main__":
    main()
