#!/usr/bin/env python3
"""
Sonda de fuentes — diagnóstico, no toca la base.

Desde el contenedor de Claude Code todas las fuentes de lotería están
bloqueadas por egress. Desde un runner de Actions sí se llega.

Historia de lo aprendido (2026-08-27):

  v1  resuloto.com para 2026-06-23 devuelve HTTP 200 con CERO bytes -> el
      hueco es de resuloto, no de Leidsa. El sorteo se hizo.
      La API de elboletoganador que documenta CLAUDE.md §7 da 404 en las 4
      rutas probadas; el sitio es un cascarón SPA de 2.5 KB.
  v2  leidsa.com trae los números embebidos en JavaScript.
  v3  Contrastado contra la base: 3_6248 (24/6) parseó EXACTA, pero 3_6244
      (20/6) dio solape 5/20. Cada página tiene ~102 secuencias de 20 números
      distintos: "la primera" es una moneda al aire. La heurística de
      update_db.py no sirve en páginas ricas.
      elboletoganador dio timeout de conexión — segunda falla seguida.
  v4  13 páginas de Leidsa bajadas para cruzarlas contra la base.

Esta v5 agrega conectate.com, que es lo que pidió Jaime, y mantiene la regla
que sacamos a golpes: NADA se da por bueno si no coincide exacto con un sorteo
que ya está en la base. Verdad de campo o no se usa.

Uso:
    python scripts/probe_fuentes.py --salida probe/
"""

import argparse
import re
import time
from html import unescape
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"}
SEC20 = re.compile(r"""(?:\b0?\d{1,2}\b[\s"']*,[\s"']*){19}\b0?\d{1,2}\b""")
DB = Path(__file__).parent.parent / "data" / "kino_2010_a_hoy_COMPLETO.xlsx"

ANCLA_ID, ANCLA_FECHA = 6248, date(2026, 6, 24)   # confirmado por <title> en v3
VEREDICTO = []
TEXTOS = {}          # nombre -> texto plano, para volcarlo al reporte


def veredicto(linea):
    VEREDICTO.append(linea)
    print(f"  >> {linea}")


def base_conocida():
    wb = load_workbook(DB, read_only=True, data_only=True)
    out = {}
    for hoja in wb.sheetnames:
        if not hoja.strip().isdigit():
            continue
        for a, b in wb[hoja].iter_rows(min_col=1, max_col=2, values_only=True):
            if not a or not b or not isinstance(b, str):
                continue
            try:
                d = date.fromisoformat(str(a).strip()[:10])
            except ValueError:
                continue
            out[d] = sorted(int(x) for x in b.replace(" ", "").split(",") if x)
    wb.close()
    return out


def candidatas(txt):
    """Todas las ventanas de 20 números distintos 1..80, SOLAPADAS.

    OJO — aquí estaba el bug de las v3/v5: se usaba re.finditer con un patrón
    de 20 números, que NO solapa. Si el sorteo real no empezaba justo en el
    borde de un match, el regex devolvía una ventana corrida un lugar: 19 de
    los 20 números buenos más un vecino. De ahí salían los "SIN MATCH (mejor
    solape 19/20)" y las "102 candidatas". No era la fuente: era el parser.

    Deslizar una ventana sobre TODOS los números en orden lo resuelve, y de
    paso sirve igual si vienen separados por coma o por espacios (conectate
    los pinta en elementos HTML sueltos, sin comas).
    """
    ns = [(m.start(), int(m.group()))
          for m in re.finditer(r"\d{1,2}", txt) if 1 <= int(m.group()) <= 80]
    out = []
    for i in range(len(ns) - 19):
        v = [n for _, n in ns[i:i + 20]]
        if len(set(v)) == 20:
            out.append((ns[i][0], sorted(v)))
    return out


def texto(html):
    """Texto plano. OJO: hay que desescapar entidades.

    Sin esto, "&#9668;&nbsp;Anterior" deja un &nbsp; literal entre el marcador
    y la palabra, y cualquier patrón anclado en "Anterior" falla. Ese fue el
    bug que hizo que el mapeo de cobertura devolviera None para los dos juegos
    incluso en fechas que sí tienen datos.
    """
    h = re.sub(r"<(script|style).*?</\1>", " ", html, flags=re.S | re.I)
    h = unescape(re.sub(r"<[^>]+>", " ", h))
    return re.sub(r"\s+", " ", h).strip()


def bajar(url, salida, nombre, timeout=30):
    try:
        r = requests.get(url, headers=UA, timeout=timeout)
    except requests.RequestException as e:
        print(f"  ERROR de red: {type(e).__name__}")
        return None
    if not r.encoding or r.encoding.lower() in ("iso-8859-1", "latin-1"):
        r.encoding = r.apparent_encoding or "utf-8"   # si no, sale "EstadÃ­sticas"
    print(f"  HTTP {r.status_code}  {len(r.content):,} bytes")
    (salida / f"{nombre}.html").write_bytes(r.content)
    return r


def cotejar(html, base, etiqueta):
    """¿Alguna secuencia de la página coincide EXACTO con un sorteo conocido?

    Esta es la única prueba que vale. Si una página trae sorteos reales, al
    menos uno tiene que dar match perfecto contra la base.
    """
    cands = candidatas(html)
    conocidos = {tuple(v): d for d, v in base.items()}
    hits = []
    for pos, v in cands:
        d = conocidos.get(tuple(v))
        if d:
            hits.append((d, pos, v))
    print(f"  {len(cands)} secuencias de 20 | {len(hits)} coinciden con la base")
    for d, pos, v in hits[:8]:
        print(f"     {d} @{pos}: {v}")
    return cands, hits


# ---------------------------------------------------------------------------
def mapear_eras_resuloto(salida, ventana=14):
    """¿Hasta dónde llega resuloto, sabiendo que la cadencia cambió?

    Jaime avisó que la frecuencia varió: Kino habría empezado con 1 o 2
    sorteos por semana y hoy es diario; Pool era miércoles y sábados. Por eso
    NO se puede muestrear días sueltos: en una época de 2 sorteos por semana,
    4 fechas al azar dan 0 de 4 y se leería como "la fuente no tiene datos".

    En su lugar se pide una VENTANA de 14 días corridos por época. Catorce días
    atrapan al menos 2 sorteos aunque la cadencia sea de 1 por semana. Y el
    conteo dentro de la ventana mide la cadencia de esa época directamente:

        14/14 sorteos -> diario        4/14 -> 2 por semana
         2/14         -> 1 por semana  0/14 -> no hay datos ahí
    """
    print(f"\n{'=' * 70}\n## COBERTURA de resuloto por épocas (ventanas de {ventana} días)\n{'=' * 70}")
    juegos = {
        "kino": ("https://www.resuloto.com/do/leid/super-kino-tv-amp.php?fecha={}", 20, 1, 80),
        "pool": ("https://www.resuloto.com/do/leid/loto-pool-amp.php?fecha={}", 5, 0, 31),
    }
    eras = [date(a, 5, 5) for a in
            (1998, 2001, 2004, 2007, 2009, 2010, 2012, 2014, 2016, 2019, 2022, 2025)]

    for juego, (url, n_esp, lo, hi) in juegos.items():
        print(f"\n### {juego}")
        pat = re.compile(r"((?:\b\d{2}\b\s+){%d})(?:◄)?\s*Anterior" % n_esp)
        for era in eras:
            hallados, dias_sem = 0, []
            for k in range(ventana):
                d = era + timedelta(days=k)
                try:
                    r = requests.get(url.format(d.isoformat()), headers=UA, timeout=20)
                except requests.RequestException:
                    continue
                if r.status_code != 200:
                    continue
                if not r.encoding or r.encoding.lower() in ("iso-8859-1", "latin-1"):
                    r.encoding = r.apparent_encoding or "utf-8"
                m = pat.search(texto(r.text))
                if m:
                    v = [int(x) for x in m.group(1).split()]
                    if len(set(v)) == n_esp and all(lo <= x <= hi for x in v):
                        hallados += 1
                        dias_sem.append("LMXJVSD"[d.weekday()])
                time.sleep(0.15)
            cadencia = hallados / ventana * 7
            barra = "#" * hallados + "." * (ventana - hallados)
            print(f"   {era}  {barra}  {hallados:>2}/{ventana}"
                  f"  ~{cadencia:.1f}/sem  dias={''.join(dias_sem)}")
            veredicto(f"{juego} {era.year}: {hallados}/{ventana} "
                      f"(~{cadencia:.1f}/semana) dias={''.join(dias_sem) or '-'}")


def sondear_pool_resuloto(salida):
    """Loto Pool en resuloto — la única fuente que ha dado datos verificados.

    Los 151 sorteos de Kino que entraron a la base salieron de aquí y pasaron
    todas las pruebas: 0 duplicados, todos con 20 números, y el del 24/6
    coincide exacto con lo que publica leidsa.com. Si tiene Kino, lo lógico es
    que tenga Pool con el mismo patrón de URL.

    Especificación de Loto Pool según leidsa vía conectate: 5 números del 00
    al 31. OJO: conectate se contradice sobre los días — su índice lo lista
    como diario y su página de Pool dice miércoles y sábados. Por eso se piden
    9 días corridos: qué días traen sorteo lo decide el dato, no el texto.
    """
    print(f"\n{'=' * 70}\n## LOTO POOL en resuloto\n{'=' * 70}")
    patrones = [
        "https://www.resuloto.com/do/leid/loto-pool-amp.php?fecha={}",
        "https://www.resuloto.com/do/leid/loto-pool.php?fecha={}",
        "https://www.resuloto.com/do/leid/pool-amp.php?fecha={}",
    ]
    # Una fecha de control donde SÍ hubo sorteo de Kino, para separar
    # "la URL no existe" de "ese día no hubo sorteo".
    ctrl = date(2026, 8, 22)
    vivo = None
    for pat in patrones:
        url = pat.format(ctrl.isoformat())
        print(f"\n### {url}")
        r = bajar(url, salida, "pool_" + re.sub(r"\W+", "_", pat)[:40])
        if r is None:
            continue
        t = texto(r.text)
        veredicto(f"pool patrón {pat.split('/')[-1].split('?')[0]}: "
                  f"HTTP {r.status_code} {len(r.content):,}B texto={len(t):,}")
        if r.status_code == 200 and len(t) > 0:
            vivo = pat
            print(f"  texto: {t[:600]}")

    if not vivo:
        veredicto("pool: NINGÚN patrón de resuloto respondió con contenido")
        return

    # Volcar la maqueta: son ~1,200 chars. Anclar el parser en "salta 3 y coge
    # 5" es justo la clase de heurística frágil que ya falló antes.
    for etq, pat in (("amp", patrones[0]), ("plano", patrones[1])):
        r = bajar(pat.format(ctrl.isoformat()), salida, f"pool_maqueta_{etq}")
        if r is None or r.status_code != 200:
            continue
        TEXTOS[f"pool_{etq}_{ctrl}"] = texto(r.text)

    veredicto(f"pool: patrón vivo -> {vivo}")
    # Nueve días corridos: así se ve solo qué días hay sorteo.
    print(f"\n### barrido de 9 días con {vivo}")
    dias = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    for i in range(9):
        d = date(2026, 8, 17) + timedelta(days=i)
        r = bajar(vivo.format(d.isoformat()), salida, f"pool_{d}", timeout=20)
        if r is None or r.status_code != 200:
            continue
        t = texto(r.text)
        ns = [int(x) for x in re.findall(r"\b\d{1,2}\b", t)]
        en_rango = [n for n in ns if 0 <= n <= 31]
        veredicto(f"pool {d} {dias[d.weekday()]}: texto={len(t):,} "
                  f"numeros={len(ns)} en_rango_00_31={en_rango[:12]}")


def sondear_conectate(salida, base):
    print(f"\n{'=' * 70}\n## CONECTATE.COM\n{'=' * 70}")
    urls = [
        ("cnt_indice", "https://loterias.conectate.com.do/leidsa/"),
        ("cnt_kino", "https://www.conectate.com.do/loterias/leidsa/super-kino-tv"),
        ("cnt_kino_sub", "https://loterias.conectate.com.do/leidsa/super-kino-tv/"),
        ("cnt_pool", "https://www.conectate.com.do/loterias/leidsa/loto-pool"),
        ("cnt_pool_sub", "https://loterias.conectate.com.do/leidsa/loto-pool/"),
        ("cnt_leidsa", "https://www.conectate.com.do/loterias/leidsa"),
    ]
    for nombre, url in urls:
        print(f"\n### {nombre}\n{url}")
        r = bajar(url, salida, nombre)
        if r is None or r.status_code != 200:
            veredicto(f"{nombre}: HTTP {r.status_code if r else 'ERROR'}")
            continue
        t = texto(r.text)
        print(f"  texto plano: {len(t):,} chars")
        TEXTOS[nombre] = t
        print(f"\n  --- texto plano de {nombre} (primeros 2200 chars) ---")
        print("  " + t[:2200])
        print("  --- fin ---\n")
        print("  [sobre HTML crudo]")
        cands, hits = cotejar(r.text, base, nombre)
        print("  [sobre texto plano]")
        cands_t, hits_t = cotejar(t, base, nombre + "/texto")
        if hits_t and not hits:
            cands, hits = cands_t, hits_t
        veredicto(f"{nombre}: {len(r.content):,}B texto={len(t):,} "
                  f"secuencias={len(cands)} MATCH_BASE={len(hits)}"
                  + (f" fechas={[str(d) for d, _, _ in hits[:5]]}" if hits else ""))

        # ¿Hay historial? Enlaces con fecha o con pinta de archivo.
        enlaces = sorted(set(re.findall(r'href="([^"]*(?:loteria|resultado|histor|fecha|20\d\d)[^"]*)"',
                                        r.text, re.I)))
        if enlaces:
            print(f"  {len(enlaces)} enlaces con pinta de historial:")
            for e in enlaces[:20]:
                print(f"     {urljoin(url, e)}")
            veredicto(f"{nombre}: {len(enlaces)} enlaces candidatos a historial")


def sondear_leidsa(salida, base, dias):
    print(f"\n{'=' * 70}\n## LEIDSA — ¿en qué posición cae la candidata correcta?\n{'=' * 70}")
    ranks = []
    for off in range(-dias, dias + 1):
        kid = ANCLA_ID + off
        url = f"https://www.leidsa.com/results/Leidsa/KinoTV/3_{kid}"
        print(f"\n### 3_{kid}")
        r = bajar(url, salida, f"leidsa_{kid}")
        if r is None or r.status_code != 200:
            continue
        mt = re.search(r"<title>\s*Leidsa KinoTV Resultados \|\s*(\d+)/(\d+)/(\d+)", r.text)
        if not mt:
            print("  sin título parseable")
            continue
        dd, mm, yy = (int(x) for x in mt.groups())
        real = date(yy, mm, dd)
        cands = candidatas(r.text)
        conocido = base.get(real)
        if not conocido:
            veredicto(f"leidsa 3_{kid}: fecha={real} NO está en la base "
                      f"({len(cands)} candidatas) <- ESTA es la que falta")
            continue
        hit = [i for i, (_, v) in enumerate(cands) if v == conocido]
        if hit:
            ranks.append(hit[0])
            pos = cands[hit[0]][0]
            ctx = re.sub(r"\s+", " ", r.text[max(0, pos - 200):pos])[-200:]
            veredicto(f"leidsa 3_{kid}: fecha={real} OK rank={hit[0]}/{len(cands)}")
            print(f"     contexto: ...{ctx}")
        else:
            mejor = max((len(set(v) & set(conocido)) for _, v in cands), default=0)
            veredicto(f"leidsa 3_{kid}: fecha={real} SIN MATCH "
                      f"(mejor solape {mejor}/20 de {len(cands)})")
    if ranks:
        veredicto(f"leidsa RANKS de la candidata correcta: {sorted(set(ranks))} "
                  + ("(CONSTANTE -> parser posicional sirve)"
                     if len(set(ranks)) == 1 else "(VARIABLE -> hace falta anclar por campo)"))


def volcado_textos():
    """Los textos planos de conectate, al reporte.

    Sin ver cómo está maquetada la página no se puede escribir el parser, y
    leerlos del log del job sale carísimo. Se deduplica por contenido: el
    dominio www y el subdominio loterias devuelven bytes idénticos.
    """
    vistos, partes = {}, ["## Textos planos (para escribir el parser)\n"]
    for nombre, t in TEXTOS.items():
        h = hash(t)
        if h in vistos:
            partes.append(f"### {nombre}\n\nIdéntico a `{vistos[h]}`.\n")
            continue
        vistos[h] = nombre
        partes.append(f"### {nombre} ({len(t):,} chars)\n\n```\n{t[:3500]}\n```\n")
    return "\n".join(partes)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--salida", default="probe")
    ap.add_argument("--dias", type=int, default=3)
    ap.add_argument("--saltar-leidsa", action="store_true")
    args = ap.parse_args()

    salida = Path(args.salida)
    salida.mkdir(parents=True, exist_ok=True)
    base = base_conocida()
    veredicto(f"base: {len(base):,} sorteos conocidos, último {max(base)}")

    mapear_eras_resuloto(salida)
    if not args.saltar_leidsa:
        sondear_leidsa(salida, base, args.dias)

    # El veredicto va al final Y se guarda aparte, para no tener que paginar
    # 300 líneas de log por tres datos.
    txt = "\n".join(f"[V] {l}" for l in VEREDICTO)
    Path("veredicto.txt").write_text(txt + "\n")
    print(f"\n{'=' * 70}\n@@ VEREDICTO @@\n{'=' * 70}\n{txt}\n{'=' * 70}")

    # Reporte al repo: paginar logs de GitHub para leer 15 líneas sale caro.
    rep = salida / "REPORTE.md"
    rep.write_text(
        "# Sonda de fuentes — último resultado\n\n"
        "Generado por `.github/workflows/probe-fuentes.yml`. Diagnóstico, no "
        "toca la base.\n\n"
        "Criterio: una fuente solo cuenta si alguna de sus secuencias de 20 "
        "números coincide EXACTO\ncon un sorteo que ya está en la base "
        "(`MATCH_BASE`). Contar bytes no prueba nada.\n\n"
        "```\n" + txt + "\n```\n\n"
        + volcado_textos())
    print(f"reporte -> {rep}")


if __name__ == "__main__":
    main()
